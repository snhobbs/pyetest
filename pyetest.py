import subprocess
import tempfile
import logging
import os
import openpyxl
import pandas
import numpy


def worksheet_to_df(worksheet):
    """
    Turn a openpyxl worksheet and create a pandas DataFrame. Drop all rows that are completely empty.
    :param openpyxl.Worksheet worksheet:
    """
    wb_data = list(worksheet.values)
    df = pandas.DataFrame(wb_data[1:], columns=wb_data[0])
    df = df.fillna(value=numpy.nan)
    df.dropna(how="all", axis=0, inplace=True)  # if all entries as NA drop them
    return df


def reload_spreadsheet(workbook):
    """
    Uses libreoffice to fill in the cached values.
    Will not update the file if the exported filename is the same.
    Filename can be any type of spreadsheet
    :param Workbook workbook:
    """
    dir_a = tempfile.mkdtemp()
    dir_b = tempfile.mkdtemp()
    fname_intermediate = os.path.join(dir_a, "pyetest.xlsx")
    fname_final = os.path.join(dir_b, "pyetest.xlsx")

    workbook.save(fname_intermediate)
    subprocess.call(
        ["libreoffice", "--convert-to", "xlsx", fname_intermediate, "--outdir", dir_b]
    )
    # logging.getLogger().info("Generated %s", outname)
    return openpyxl.open(fname_final, data_only=True)


def get_column_number_dict(workbook):
    """
    Generates a lookup table from column name to column number
    :param Workbook workbook: template object
    """
    column_names = {}
    for (column,) in workbook.active.iter_cols(min_row=1, max_row=1, values_only=False):
        column_names[column.value] = column.column - 1  # _letter
    return column_names


def compile_template(workbook, data):
    """
    Fills a template with values from a seperate dataframe
    :param Workbook workbook: Template to be filled, field names need to be correct
    :param [dict, DataFrame] data: Table of names and values matching the names in the template
    """
    starting_row = 1
    column_names = get_column_number_dict(workbook)
    for row in workbook.active.iter_rows(min_row=starting_row + 1, values_only=False):
        if row[column_names["type"]].value == "measured":
            column = column_names["value"] + 1  # counts from 1
            coordinate = workbook.active.cell(row[0].row, column=column).coordinate
            name = row[column_names["name"]].value
            if name is None:
                continue
            try:
                value = data[data["name"] == name].iloc[0]["value"]
            except IndexError:
                logging.error(
                    "Entry %s not in data, either remove from tests or add to data",
                    name,
                )
                raise
            workbook.active[coordinate] = value
    return workbook


def compile_test_template(workbook, data):
    """
    Takes a test template and fills in the values field.
    Essentially identical to filling the data template without the measured filtering
    :param Workbook workbook: Template to be filled, field names need to be correct
    :param [dict, DataFrame] data: Table of names and values matching the names in the template
    """
    starting_row = 1
    column_names = get_column_number_dict(workbook)
    for row in workbook.active.iter_rows(min_row=starting_row + 1, values_only=False):
        column = column_names["value"] + 1  # counts from 1
        coordinate = workbook.active.cell(row[0].row, column=column).coordinate
        name = row[column_names["name"]].value
        if name is None:
            continue
        try:
            value = data[data["name"] == name].iloc[0]["value"]
        except IndexError:
            logging.error(
                "Entry %s not in data, either remove from tests or add to data", name
            )
            raise
        workbook.active[coordinate] = value
    return workbook


def pyetest_run(data_template, test_template, data_df):
    """
    Fill the data workbook, reload it, turn the workbook into a dataframe, fill the tests with values and reload that. The test workbook can be turned into a dataframe and pass/fail test run.
    :param Workbook data_template: Data template to be filled, field names need to be correct
    :param Workbook test_template: Test template to be filled, field names need to be correct
    :param [dict, DataFrame] data_df: Table of names and values matching the names in the template
    """
    compiled_wb = compile_template(data_template, data_df)
    data_workbook = reload_spreadsheet(compiled_wb)
    compiled_test_wb = compile_test_template(
        test_template, worksheet_to_df(data_workbook.active)
    )
    test_workbook = reload_spreadsheet(compiled_test_wb)
    return data_workbook, test_workbook
